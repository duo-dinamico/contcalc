#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Containment Calculation Sheet
menuClass.py v1.0
#odetojoy
Copyright 2020, The JJ duo
"""

import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from tkinter.filedialog import askopenfilename, asksaveasfilename
from tabClass import MyTab, MyCable
from widgetsClass import CreateLabel, CreateEntry, CreateButton
import platform
import json
from openpyxl import load_workbook
from openpyxl.drawing.image import Image


class Menu(tk.Menu):
    """Class that create the menu.

    INPUT: self
    OUTPUT: no output
    """
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent

        # The following 14 lines of code create our menu and associate commands to the options
        self.filemenu = tk.Menu(self, tearoff=0)
        self.editmenu = tk.Menu(self, tearoff=0)
        self.helpmenu = tk.Menu(self, tearoff=0)
        self.add_cascade(label='File', menu=self.filemenu)
        self.add_cascade(label='Edit', menu=self.editmenu)
        self.add_cascade(label='Help', menu=self.helpmenu)
        self.filemenu.add_command(label='Open', command=self.open_file)
        self.filemenu.add_command(label='Save', command=self.save_file)
        self.filemenu.entryconfigure(1, state='disabled')
        self.filemenu.add_command(label='Save as...', command=self.saveas_file)
        self.filemenu.add_separator()
        self.filemenu.add_command(label='Export to Excel', command=self.export_excel)
        self.filemenu.add_separator()
        self.filemenu.add_command(label='Exit', command=self.confirm_exit)
        self.editmenu.add_command(label='Add Section...', command=lambda: self.popupWindow('add'))
        self.editmenu.add_command(label='Delete Section...', command=lambda: self.popupWindow('del'))
        self.helpmenu.add_command(label='About', command=self.about_menu)

        self.filename = ''

    def json_access(self, mode, dialog):
        """Method to load and dump data from a json.
        INPUT: JSON file with entries and tabs
        OUTPUT: no output
        """
        self.mode = mode # the mode will determine if we're doing Write or Read
        self.dialog = dialog  # the dialog will determine if we're doing Save as Filename or Open Filename

        if self.dialog == asksaveasfilename or self.dialog == askopenfilename:  # if dialog asks for dialog box, give it to the user
            self.filepath = self.dialog(
                filetypes=[('JSON File', '*.txt'), ('All files', '*.*')]
            )
            if not self.filepath:
                return

        if '.txt' not in self.filepath and self.dialog == asksaveasfilename:
            self.filepath = self.filepath + '.txt'
        else:
            pass
        with open(self.filepath, self.mode, encoding='utf-8') as json_file:  # access the json file as Write or Read mode
            if self.dialog == askopenfilename:  # if open file, access the file and save it on a variable
                data = json.load(json_file)
                self.filemenu.entryconfigure(1, state='normal')  # set menu save to normal
            else:
                json.dump(self.to_save(), json_file, ensure_ascii=False, indent=4, separators=(',', ': ')) # otherwise just do a dump into the file
                data = ''  # set data to empty just for error handling
        return data  # return the data to be used when opening a file

    def open_file(self):
        """Method for opening files.
        INPUT: JSON file with entries and tabs
        OUTPUT: no output
        """
        data = self.json_access('r', askopenfilename)  # convert the data into a variable inside the open
        try:
            if data['Project Info']['Software Version'] == self.parent.contcalc_version:
                for v in self.parent.nb.tabs_list.values():  # cycle through open tabs and destroy them
                    v.destroy()
                self.parent.nb.create_general_tab('Main Page')  # generate a clear main page
                for i, myline in zip(self.parent.nb.lst_entries, data['Project Info'].values()):  # cycle through the list of tk entries to fill them with the values from data
                    i.set(myline)
                for d in data['Project Tabs']:
                    new_tab = MyTab(self.parent.nb, d['Tab_name'])
                    if d['Installation type'] == 'Custom Spacing':
                        new_tab.common_spacing_entry.config(state='normal')
                    else:
                        pass
                    new_tab.common_install_combobox.set(d['Installation type'])
                    new_tab.common_spacing_var.set(d['Custom Spacing'])
                    new_tab.common_cont_combobox.set(d['Containment Type'])
                    new_tab.common_spare_var.set(d['Spare Capacity'])
                    for z in d['Cables']:
                        cable = MyCable(z['Reference'], z['Type'], z['Number of cables'], z['CSA'], z['No Parallels'], z['CPC CSA'])
                        new_tab.cable_list.append(cable)
                        new_tab.populate_list()
                        new_tab.print_result()
                        new_tab.clear_cable()
                    self.dict = {d['Tab_name']: new_tab}
                    self.parent.nb.tabs_list.update(self.dict)

                self.filename_state('normal')  # function to set title and save menu to enabled
        except KeyError:
            messagebox.showwarning(title='Error', message='Incorrect file. Please chose a correct save file.')

    def filename_state(self, state):
        self.filename = self.filepath.lower().split('/')[-1].replace('.txt', '')  # clean filepath to get filename only
        self.filename = self.filename.title()
        self.parent.title(f'Containment Calculation Sheet - {self.filename}')  # set the title of the window to the filename
        self.filemenu.entryconfigure(1, state=state)  # set save menu to enabled

    def save_file(self):
        """Method for saving files.
        INPUT: self
        OUTPUT: JSON file with entries and tabs
        """
        try:
            if self.filename: # if the software has a name, should be able to save it to the filepath. if not, silently pass
                self.json_access('w', '')
        except AttributeError:
            print('failed')

    def saveas_file(self):
        """Method for saving files as.
        INPUT: self
        OUTPUT: JSON file with entries and tabs
        """
        self.json_access('w', asksaveasfilename)
        self.filename_state('normal')  # function to set title and save menu to enabled

    def export_excel(self):
        """ Method to export data to Excel. """

        # Load template file
        twb = load_workbook('TemplateContCalc.xlsx')

        # Get data
        data = self.parent.nb.get_notebook_dict(True)

        # Get image file
        try:
            img = Image('images/ArupLogo.gif')
        except FileNotFoundError:
            messagebox.showwarning(title='Error', message='Image not found.')
            print(f'Image not found.')
            return

        # Insert image in Workbook
        twb['Main Page'].add_image(img, 'A1')

        # Insert data in project info
        twb['Main Page']['D9'] = data['Project Info']['Job Title']
        twb['Main Page']['D11'] = data['Project Info']['Job Number']
        twb['Main Page']['H11'] = data['Project Info']['Designer']
        twb['Main Page']['D13'] = data['Project Info']['Date']
        twb['Main Page']['H13'] = data['Project Info']['Revision']

        # Copy from Temp_Tab, and insert data
        for dict in data['Project Tabs']:
            sheet = twb.copy_worksheet(twb['TrayTemplate'])
            sheet.title = dict['Tab_name']
            sheet['D5'] = dict['Installation type']
            sheet['I5'] = dict['Custom Spacing']
            sheet['D7'] = dict['Containment Type']
            sheet['D9'] = dict['Spare Capacity']
            sheet['I9'] = dict['Tab_name']

            # Cables
            my_row = 12
            for cable_dict in dict['Cables']:
                sheet.cell(row=my_row, column=2, value=cable_dict['Reference'])
                sheet.cell(row=my_row, column=3, value=cable_dict['Type'])
                sheet.cell(row=my_row, column=5, value=cable_dict['Number of cables'])
                sheet.cell(row=my_row, column=6, value=cable_dict['CSA'])
                sheet.cell(row=my_row, column=7, value=cable_dict['No Parallels'])
                sheet.cell(row=my_row, column=9, value=cable_dict['CPC CSA'])
                sheet.cell(row=my_row, column=11, value=cable_dict['Diameter'])
                my_row += 1

            # Results
            sheet['I34'] = dict['Results']['Total diameter']
            sheet['I36'] = dict['Results']['Total with spare']
            sheet['I38'] = dict['Results']['Containment size']

        # Delete Temp_Tab
        twb.remove_sheet(twb['TrayTemplate'])

        # Save Workbook to file
        dest_filename = tk.filedialog.asksaveasfilename(
            initialdir=".",
            initialfile=self.filename,
            title="Select file to export",
            defaultextension="*.xlsx",
            filetypes=(("Excel files", "*.xlsx"), ("all files", "*.*"))
            )

        # Check if not canceled
        if dest_filename != '':
            twb.save(dest_filename)
        else:
            print("Export canceled.")

        # Just to check structure
        # print(json.dumps(data, indent=4, separators=(',', ': ')))

    def to_save(self):
        to_save = self.parent.nb.get_notebook_dict(False)
        return to_save

    def about_menu(self):
        """Method for versions.
        INPUT: self
        OUTPUT: no output
        """
        py_version_compact = platform.python_version()
        tkinter_version = tk.TkVersion
        messagebox.showinfo(title='About', message=(f'This tools versions is: {self.parent.contcalc_version}\n\nYour python version is: {py_version_compact}\n\nYour tkinter version is: {tkinter_version}'))

    def confirm_exit(self):
        """Method to get warning box about exiting.
        INPUT: no input
        OUTPUT: no output
        """
        self.confirm_exit_dialog = messagebox.askyesnocancel(title='Save on close', message='Do you want to save before quiting?', default=tk.messagebox.YES, icon='question')

        try:
            if self.confirm_exit_dialog and self.filename:
                self.save_file()
                self.parent.quit()
            elif self.confirm_exit_dialog:
                self.saveas_file()
                self.parent.quit()
            elif self.confirm_exit_dialog is None:
                return
            else:
                self.parent.quit()
        except AttributeError:
            if self.confirm_exit_dialog:
                self.saveas_file()
                self.parent.quit()
            elif self.confirm_exit_dialog is None:
                return
            else:
                self.parent.quit()

    def popupWindow(self, mode):
        self.top = tk.Toplevel(self.parent)
        self.top.focus()
        self.top.title('Section Name')
        self.top.wm_geometry('225x125+300+300')
        self.top.transient(self.parent)
        self.top.resizable(False, False)
        self.pu_frame = ttk.Frame(self.top)
        self.pu_frame.pack()
        self.pu_label = CreateLabel(self.pu_frame, text='Please enter name of section', image='', background=None, height=2, width=25, row=0, column=0, colspan=1, sticky='EW')
        self.empty = CreateLabel(self.pu_frame, text='', image='', background=None, height=2, width=25, row=2, column=0, colspan=1, sticky='EW')
        self.pu_button = CreateButton(self.pu_frame, text='', height=1, width=10, command=self.add_tab, row=3, column=0, colspan=1, sticky='')
        if mode == 'add':
            self.pu_entry = CreateEntry(self.pu_frame, background='white', row=1, column=0, colspan=1, sticky='')
            self.pu_entry.focus_set()
            self.pu_button.config(text='Add', command=self.add_tab)
        else:
            if len(self.parent.nb.tabs_list.keys()) > 1:
                self.delete_tab_combobox = ttk.Combobox(self.pu_frame, values=list(self.parent.nb.tabs_list.keys())[1:], state='readonly')
                self.delete_tab_combobox.grid(row=1, column=0)
                self.pu_button.config(text='Delete', command=self.del_tab)
            else:
                self.top.destroy()
                messagebox.showwarning(title='Error', message='No sections to remove.')

    def add_tab(self):
        popup_value = self.pu_entry.get().strip()

        for k in self.parent.nb.tabs_list:
            if k == popup_value:
                messagebox.showwarning(title='Error', message='That section name already exist.')
                return
        if popup_value != '':
            self.parent.nb.tab_create(popup_value)
            self.top.destroy()
        else:
            self.empty.config(text='You must insert a section name.')

    def del_tab(self):
        for k, v in self.parent.nb.tabs_list.items():
            if k == self.delete_tab_combobox.get():
                self.top.destroy()
                v.delete_this_tab(False)
                return
